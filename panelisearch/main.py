import csv
import json
import re
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import Response

BASE_DIR = Path(__file__).resolve().parent
DATABASE_DIR = BASE_DIR / "database"
PORT = 8080
MAX_RESULTS = 50
SUPPORTED_SUFFIXES = {".csv", ".txt", ".json", ".tsv", ".xlsx", ".xlsm"}

COLUMN_MAP = {
    "name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "ad": "first_name",
    "isim": "first_name",
    "surname": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "soyad": "last_name",
    "soyisim": "last_name",
    "phone": "phone",
    "phone number": "phone",
    "telephone": "phone",
    "gsm": "phone",
    "mobile": "phone",
    "tel": "phone",
    "telefon": "phone",
    "cep": "phone",
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "eposta": "email",
    "city": "city",
    "country": "country",
    "tc": "identity_number",
    "tc kimlik": "identity_number",
    "identity_number": "identity_number",
    "id": "identity_number",
}


def normalize(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def phone_digits(value: str) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def clean_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def clean_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def line_matches(query: str, text: str) -> bool:
    needle = normalize(query)
    haystack = normalize(text)
    if needle in haystack:
        return True

    query_digits = phone_digits(query)
    if len(query_digits) >= 4 and query_digits in phone_digits(text):
        return True
    return False


def map_row(raw: dict) -> dict:
    record = {
        "first_name": "",
        "last_name": "",
        "phone": "",
        "email": "",
        "identity_number": "",
        "city": "",
        "country": "",
        "notes": "",
        "extra": {},
    }

    for key, value in raw.items():
        header = clean_header(key)
        text = clean_value(value)
        if not text or text.lower() in {"null", "none", "nan"}:
            continue

        field = COLUMN_MAP.get(header)
        if field in record and field != "extra":
            record[field] = text
        elif header:
            record["extra"][header] = text

    return record


def record_to_result(record: dict) -> dict | None:
    first = record.get("first_name", "")
    last = record.get("last_name", "")
    full_name = f"{first} {last}".strip()
    if not full_name and record.get("extra", {}).get("name"):
        full_name = str(record["extra"]["name"])

    result = {
        "full_name": full_name,
        "first_name": first,
        "last_name": last,
        "phone": record.get("phone", ""),
        "email": record.get("email", ""),
        "identity_number": record.get("identity_number", ""),
        "city": record.get("city", ""),
        "country": record.get("country", ""),
        "notes": record.get("notes", ""),
    }

    extra = record.get("extra") or {}
    for key, value in extra.items():
        lowered = clean_header(key)
        if lowered in {"name", "full_name", "fullname"} and not result["full_name"]:
            result["full_name"] = clean_value(value)
        elif lowered in COLUMN_MAP and not result.get(COLUMN_MAP[lowered]):
            result[COLUMN_MAP[lowered]] = clean_value(value)

    if not any(result.values()) and not extra:
        return None

    if not result["full_name"]:
        result["full_name"] = result["email"] or result["phone"] or ""

    cleaned = {key: value for key, value in result.items() if value}
    if extra and not any(cleaned.values()):
        cleaned["details"] = " | ".join(f"{k}: {v}" for k, v in extra.items())
    elif extra:
        cleaned["details"] = " | ".join(f"{k}: {v}" for k, v in extra.items() if k not in COLUMN_MAP)

    return cleaned or None


def record_matches(query: str, record: dict) -> bool:
    parts = [
        record.get("first_name", ""),
        record.get("last_name", ""),
        record.get("phone", ""),
        record.get("email", ""),
        record.get("identity_number", ""),
        record.get("city", ""),
        record.get("country", ""),
        record.get("notes", ""),
    ]
    parts.extend(str(value) for value in (record.get("extra") or {}).values())
    return line_matches(query, " ".join(part for part in parts if part))


def search_csv(path: Path, query: str) -> list[dict]:
    matches: list[dict] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            record = map_row(row)
            if record_matches(query, record):
                result = record_to_result(record)
                if result:
                    matches.append(result)
            if len(matches) >= MAX_RESULTS:
                break
    return matches


def search_txt(path: Path, query: str) -> list[dict]:
    matches: list[dict] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            text = line.strip()
            if text and line_matches(query, text):
                result = record_to_result({"notes": text})
                if result:
                    matches.append(result)
            if len(matches) >= MAX_RESULTS:
                break
    return matches


def flatten_json(value, prefix: str = "") -> list[str]:
    parts: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            parts.extend(flatten_json(item, next_prefix))
    elif isinstance(value, list):
        for index, item in enumerate(value):
            next_prefix = f"{prefix}[{index}]"
            parts.extend(flatten_json(item, next_prefix))
    else:
        parts.append(f"{prefix}: {value}" if prefix else str(value))
    return parts


def search_json(path: Path, query: str) -> list[dict]:
    matches: list[dict] = []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return matches

    items = payload if isinstance(payload, list) else [payload]
    for item in items:
        if isinstance(item, dict):
            record = map_row(item)
        else:
            record = {"notes": str(item), "extra": {}}
        if record_matches(query, record) or line_matches(query, " ".join(flatten_json(item))):
            result = record_to_result(record if isinstance(item, dict) else {"notes": str(item), "extra": {}})
            if result:
                matches.append(result)
        if len(matches) >= MAX_RESULTS:
            break
    return matches


def iter_xlsx_records(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    headers: list[str] = []
    for row in rows:
        headers = [clean_header(cell) for cell in row]
        if any(headers):
            break

    for row in rows:
        raw = {}
        for index, header in enumerate(headers):
            if not header or index >= len(row):
                continue
            raw[header] = row[index]
        if any(clean_value(value) for value in raw.values()):
            yield map_row(raw)

    workbook.close()


def search_xlsx(path: Path, query: str) -> list[dict]:
    matches: list[dict] = []
    for record in iter_xlsx_records(path):
        if record_matches(query, record):
            result = record_to_result(record)
            if result:
                matches.append(result)
        if len(matches) >= MAX_RESULTS:
            break
    return matches


def database_files() -> list[Path]:
    if not DATABASE_DIR.exists():
        return []

    files: list[Path] = []
    for path in sorted(DATABASE_DIR.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("."):
            continue
        if path.suffix.lower() in SUPPORTED_SUFFIXES:
            files.append(path)
    return files


def search_files(query: str) -> list[dict]:
    if not query.strip():
        raise ValueError("Send a name, phone, email, or keyword to search")

    results: list[dict] = []
    seen: set[str] = set()

    for path in database_files():
        suffix = path.suffix.lower()
        if suffix == ".json":
            batch = search_json(path, query)
        elif suffix in {".csv", ".tsv"}:
            batch = search_csv(path, query)
        elif suffix == ".txt":
            batch = search_txt(path, query)
        elif suffix in {".xlsx", ".xlsm"}:
            batch = search_xlsx(path, query)
        else:
            continue

        for item in batch:
            key = json.dumps(item, sort_keys=True, ensure_ascii=False)
            if key in seen:
                continue
            seen.add(key)
            results.append(item)
            if len(results) >= MAX_RESULTS:
                return results

    return results


def raw_json(**data) -> Response:
    return Response(
        content=json.dumps({"ok": True, **data}, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
    )


@asynccontextmanager
async def lifespan(_app: FastAPI):
    from bot import start_telegram_bot_thread

    start_telegram_bot_thread()
    yield


app = FastAPI(title="PaneliSearch API", docs_url=None, redoc_url=None, openapi_url=None, lifespan=lifespan)


@app.get("/api")
def api(q: str | None = Query(default=None, description="Search query")) -> Response:
    if not q or not q.strip():
        return raw_json(ready=True, bot="panelisearch")

    try:
        results = search_files(q)
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error

    return raw_json(
        query=q.strip(),
        found=len(results),
        results=results,
    )


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "bot": "panelisearch"}


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=PORT)
