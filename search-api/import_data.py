#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from database import connect, ensure_database, insert_person, normalize_email, normalize_phone, normalize_text

COLUMN_MAP = {
    "name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "surname": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "phone": "phone",
    "phone number": "phone",
    "phone_number": "phone",
    "telephone": "phone",
    "email": "email",
    "e-mail": "email",
    "e-mail contact": "email",
    "e-mail contact ": "email",
    "mail": "email",
    "identity_number": "identity_number",
    "identity number": "identity_number",
    "id number": "identity_number",
    "tc": "identity_number",
    "city": "city",
    "country": "country",
    "source": "source",
    "notes": "notes",
    "ip": "notes",
}


def clean_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def map_row(raw: dict) -> dict[str, str]:
    mapped: dict[str, str] = {
        "first_name": "",
        "last_name": "",
        "phone": "",
        "email": "",
        "identity_number": "",
        "city": "",
        "country": "",
        "source": "",
        "notes": "",
    }

    for key, value in raw.items():
        header = clean_header(key)
        field = COLUMN_MAP.get(header)
        if not field:
            continue
        text = str(value or "").strip()
        if text.lower() in {"x", "none", "null", "nan"}:
            continue
        if field == "notes" and mapped["notes"]:
            mapped["notes"] = f"{mapped['notes']} | {text}"
        else:
            mapped[field] = text

    return mapped


def row_is_valid(record: dict[str, str]) -> bool:
    return any([
        record["first_name"],
        record["last_name"],
        record["phone"],
        record["email"],
        record["identity_number"],
    ])


def read_csv_rows(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [map_row(row) for row in reader]


def read_xlsx_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_header(cell) for cell in next(rows, [])]

    mapped_rows = []
    for row in rows:
        raw = {}
        for index, header in enumerate(headers):
            if not header or index >= len(row):
                continue
            raw[header] = row[index]
        mapped_rows.append(map_row(raw))

    workbook.close()
    return mapped_rows


def import_file(path: Path, replace: bool = False) -> int:
    ensure_database()

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        records = read_xlsx_rows(path)
    else:
        records = read_csv_rows(path)

    valid_records = [record for record in records if row_is_valid(record)]

    with connect() as conn:
        if replace:
            conn.execute("DELETE FROM people")

        batch = 0
        for record in valid_records:
            insert_person(conn, record)
            batch += 1
            if batch % 1000 == 0:
                print(f"Imported {batch} records...")

    return len(valid_records)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import people records into SQLite")
    parser.add_argument("file", type=Path, help="CSV, TSV, or XLSX file")
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Clear existing records before import",
    )
    args = parser.parse_args()

    if not args.file.exists():
        raise SystemExit(f"File not found: {args.file}")

    count = import_file(args.file, replace=args.replace)
    print(f"Imported {count} records from {args.file.name}.")


if __name__ == "__main__":
    main()
