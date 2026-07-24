import csv
import itertools
import json
import os
import re
import shutil
import sqlite3
import subprocess
import threading
import time
import tempfile
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import Response

BASE_DIR = Path(__file__).resolve().parent
DATABASE_DIR = BASE_DIR / "databases"
INDEX_DB = BASE_DIR / ".search_index.db"
SCHEMA_VERSION = 9
PORT = 8080
API_VERSION = "2026-07-24-telegram-fix2"
CREDIT = "api made by Ami.192 on signal"
AUTO_REBUILD = os.environ.get("AUTO_REBUILD", "0") == "1"
API_USAGE = {
    "status": "/api",
    "stats": "/api/stats",
    "rebuild": "/api/rebuild",
    "reimport": "/api/reimport",
    "name": "/api?q=Mootje bicep",
    "phone": "/api?q=905544784243",
    "email": "/api?q=email@example.com",
    "id": "/api?q=12345678901",
}
INDEX_LOCK = threading.Lock()
INDEX_READY = threading.Event()
INDEX_BUILDING = threading.Event()
INDEX_ERROR: str | None = None
BATCH_SIZE = 20000
PROGRESS_EVERY = 50000
LARGE_FILE_BYTES = 100_000_000
DATA_SUFFIXES = {".xlsx", ".xlsm", ".csv", ".tsv", ".txt", ".db", ".sql"}
ARCHIVE_SUFFIXES = {".7z"}
IMPORT_FIELDS = (
    "first_name", "last_name", "phone", "email",
    "identity_number", "city", "country", "notes",
)
SOURCE_META_KEY = "indexed_sources"
TABLE_NAME_HINTS = {
    "people", "person", "persons", "users", "user", "customers", "customer",
    "citizens", "citizen", "members", "member", "contacts", "contact",
    "records", "record", "data", "clients", "client", "kisi", "kisiler",
}

PG_COPY_RE = re.compile(
    r'^COPY\s+(?:[\w.]+\.)?(?:"([^"]+)"|(\w+))\s*\(([^)]+)\)\s+FROM\s+stdin\s*;?\s*$',
    re.IGNORECASE,
)
PG_NULLS = {"\\N", "<NULL>", "NULL", "null", "None"}
PG_PROGRESS_EVERY = 500_000

INSERT_SQL = """
INSERT INTO people (
    first_name, last_name, phone, email, identity_number,
    city, country, notes, extra_json,
    first_name_n, last_name_n, phone_n, email_n, identity_number_n
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


@asynccontextmanager
async def lifespan(_app: FastAPI):
    thread = threading.Thread(target=build_index_background, daemon=True)
    thread.start()
    from telegram_bot import start_telegram_bot_thread

    start_telegram_bot_thread()
    yield


app = FastAPI(docs_url=None, redoc_url=None, openapi_url=None, lifespan=lifespan)

COLUMN_MAP = {
    "name": "first_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "ad": "first_name",
    "first": "first_name",
    "isim": "first_name",
    "musteri adi": "first_name",
    "surname": "last_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "soyad": "last_name",
    "last": "last_name",
    "soyisim": "last_name",
    "soyadi": "last_name",
    "phone": "phone",
    "phone number": "phone",
    "phone_number": "phone",
    "telephone": "phone",
    "gsm": "phone",
    "mobile": "phone",
    "cell": "phone",
    "tel": "phone",
    "telefon": "phone",
    "cep": "phone",
    "cep telefonu": "phone",
    "cep no": "phone",
    "gsm no": "phone",
    "mobile phone": "phone",
    "contact phone": "phone",
    "phone no": "phone",
    "telefon no": "phone",
    "telefon numarasi": "phone",
    "tel no": "phone",
    "numara": "phone",
    "email": "email",
    "e-mail": "email",
    "e-mail contact": "email",
    "e-mail contact ": "email",
    "e-posta": "email",
    "eposta": "email",
    "mail": "email",
    "mail adresi": "email",
    "email address": "email",
    "identity_number": "identity_number",
    "identity number": "identity_number",
    "id number": "identity_number",
    "tc": "identity_number",
    "tc kimlik": "identity_number",
    "tc kimlik no": "identity_number",
    "tc_kimlik": "identity_number",
    "tc_kimlik_no": "identity_number",
    "tcno": "identity_number",
    "tc no": "identity_number",
    "kimlik": "identity_number",
    "kimlik_no": "identity_number",
    "kimlik no": "identity_number",
    "kimlik numarasi": "identity_number",
    "id no": "identity_number",
    "national id": "identity_number",
    "national_identifier": "identity_number",
    "id": "identity_number",
    "birth_city": "city",
    "address_city": "city",
    "dogum yeri": "city",
    "dogum_yeri": "city",
    "il": "city",
    "ilce": "notes",
    "gsm_no": "phone",
    "cep_no": "phone",
    "city": "city",
    "country": "country",
    "source": "source",
    "notes": "notes",
    "ip": "notes",
}


def clean_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def clean_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value) >= 1e9:
            return str(int(round(value)))
        if value.is_integer():
            return str(int(value))
        return f"{value:.0f}".strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def looks_like_tc(digits: str) -> bool:
    return (
        len(digits) == 11
        and not digits.startswith("0")
        and not digits.startswith("90")
        and not digits.startswith("5")
    )


def looks_like_phone_digits(digits: str) -> bool:
    if len(digits) < 7:
        return False
    if digits.startswith("90"):
        return len(digits) >= 11
    if digits.startswith("0"):
        return len(digits) >= 10
    if digits.startswith("5"):
        return len(digits) >= 10
    return len(digits) >= 10


def phone_digits(value: str | None) -> str:
    return re.sub(r"\D+", "", clean_value(value))


def norm_phone(value: str | None) -> str:
    digits = phone_digits(value)
    if not digits:
        return ""

    while digits.startswith("90") and len(digits) > 10:
        digits = digits[2:]
    while digits.startswith("0") and len(digits) > 10:
        digits = digits[1:]

    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def phone_keys(value: str | None) -> set[str]:
    digits = phone_digits(value)
    if not digits:
        return set()

    keys = {digits}
    core = norm_phone(value)
    if core:
        keys.add(core)
        keys.add("0" + core)
        keys.add("90" + core)

    if digits.startswith("90") and len(digits) > 2:
        rest = digits[2:]
        keys.add(rest)
        if not rest.startswith("0"):
            keys.add("0" + rest)

    if digits.startswith("0") and len(digits) > 1:
        keys.add(digits[1:])
        keys.add("90" + digits[1:])

    if len(digits) >= 10:
        keys.add(digits[-10:])

    return {key for key in keys if len(key) >= 7}


def phone_search_variants(value: str | None) -> set[str]:
    variants = phone_keys(value)
    core = norm_phone(value)
    if core:
        variants.add(core)
        for size in (10, 9, 8, 7):
            if len(core) >= size:
                variants.add(core[-size:])
    return {item for item in variants if len(item) >= 7}


def collect_phone_keys(record: dict) -> str:
    keys: set[str] = set()

    for field in ("phone", "notes", "identity_number"):
        keys.update(phone_keys(str(record.get(field, ""))))

    for value in record.get("extra", {}).values():
        text = clean_value(value)
        digits = phone_digits(text)
        if looks_like_phone_digits(digits):
            keys.update(phone_keys(text))

    if not keys:
        for value in record.get("extra", {}).values():
            keys.update(phone_keys(str(value)))

    return "|".join(sorted(keys, key=len, reverse=True))


PHONE_DIGITS_SQL = (
    "replace(replace(replace(replace(replace(replace(phone, ' ', ''), '-', ''), '+', ''), '(', ''), ')', ''), '.', '')"
)
IDENTITY_DIGITS_SQL = (
    "replace(replace(replace(identity_number, ' ', ''), '-', ''), '.', '')"
)


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def norm_email(value: str | None) -> str:
    return str(value or "").strip().casefold()


def map_row(raw: dict) -> dict:
    mapped = {
        "first_name": "",
        "last_name": "",
        "phone": "",
        "email": "",
        "identity_number": "",
        "city": "",
        "country": "",
        "notes": "",
        "extra": {},
    }

    for key, value in raw.items():
        header = clean_header(key)
        text = clean_value(value)
        if text.lower() in {"x", "none", "null", "nan", ""}:
            continue

        field = COLUMN_MAP.get(header)
        if field == "phone":
            text = phone_digits(text) or text
        if field == "notes" and mapped["notes"]:
            mapped["notes"] = f"{mapped['notes']} | {text}"
        elif field:
            mapped[field] = text
        elif header:
            mapped["extra"][header] = text

    if not mapped["phone"]:
        for value in raw.values():
            digits = phone_digits(clean_value(value))
            if looks_like_tc(digits):
                continue
            if looks_like_phone_digits(digits):
                mapped["phone"] = digits
                break

    return mapped


def row_is_valid(record: dict) -> bool:
    return any([
        record["first_name"],
        record["last_name"],
        record["phone"],
        record["email"],
        record["identity_number"],
        record["city"],
        record["country"],
        record["notes"],
        record["extra"],
    ])


HEADER_WORDS = {
    "name", "surname", "phone", "email", "ad", "soyad", "telefon", "mail",
    "isim", "soyisim", "gsm", "e-mail", "eposta", "numara", "firstname", "lastname",
}


def split_line(line: str, delimiter: str) -> list[str]:
    if delimiter == "space":
        return [part.strip() for part in re.split(r"\s{2,}", line.strip()) if part.strip()]
    return [part.strip() for part in line.split(delimiter)]


def line_is_header(parts: list[str]) -> bool:
    if len(parts) >= 2:
        text = clean_header(" ".join(parts))
        return any(word in HEADER_WORDS for word in text.split())

    single = clean_header(parts[0] if parts else "")
    for delimiter in ["|", ";", "\t", ","]:
        if delimiter in parts[0]:
            return line_is_header(split_line(parts[0], "|" if delimiter == "|" else delimiter))
    return any(word in HEADER_WORDS for word in single.split())


def record_quality(record: dict) -> int:
    score = 0
    if record["phone"]:
        score += 3
    if record["email"]:
        score += 3
    if record["first_name"]:
        score += 1
    if record["last_name"]:
        score += 1
    if record["identity_number"]:
        score += 1
    return score


def map_row_positional(parts: list[str]) -> dict:
    raw: dict = {}
    if len(parts) >= 5:
        raw = {
            "name": parts[0],
            "surname": parts[1],
            "phone number": parts[2],
            "e-mail": parts[3],
            "identity_number": parts[4],
        }
    elif len(parts) >= 4:
        raw = {
            "name": parts[0],
            "surname": parts[1],
            "phone number": parts[2],
            "e-mail": parts[3],
        }
    elif len(parts) == 3:
        if "@" in parts[2]:
            raw = {"name": parts[0], "surname": parts[1], "e-mail": parts[2]}
        else:
            raw = {"name": parts[0], "surname": parts[1], "phone number": parts[2]}
    elif len(parts) == 2:
        if "@" in parts[1]:
            raw = {"name": parts[0], "e-mail": parts[1]}
        else:
            raw = {"name": parts[0], "phone number": parts[1]}
    elif len(parts) == 1:
        raw = {"name": parts[0]}
    return map_row(raw)


def choose_best_delimiter(lines: list[str], suffix: str = ".txt") -> str:
    if suffix == ".tsv":
        return "\t"

    options = ["\t", "|", ";", ",", "space"]
    sample_lines = lines[:100]
    best = "\t"
    best_score = -1

    for delimiter in options:
        column_counts = [len(split_line(line, delimiter)) for line in sample_lines[:20]]
        avg_columns = sum(column_counts) / max(len(column_counts), 1)
        if avg_columns < 2:
            continue

        score = 0
        first_parts = split_line(sample_lines[0], delimiter)
        has_header = line_is_header(first_parts)
        headers = [clean_header(part) for part in first_parts] if has_header else []
        data_lines = sample_lines[1:] if has_header else sample_lines

        for line in data_lines:
            parts = split_line(line, delimiter)
            if len(parts) < 2:
                continue
            if headers:
                raw = {
                    headers[index]: parts[index] if index < len(parts) else ""
                    for index in range(len(headers))
                }
                record = map_row(raw)
            else:
                record = map_row_positional(parts)
            score += record_quality(record)

        if score > best_score:
            best_score = score
            best = delimiter

    if best_score <= 0:
        return detect_delimiter("\n".join(lines[:20]), suffix)
    return best


def detect_delimiter(sample: str, suffix: str = ".csv") -> str:
    if suffix == ".tsv":
        return "\t"
    options = ["\t", ",", ";", "|"]
    best = max(options, key=lambda item: sample.count(item))
    if sample.count(best) > 0:
        return best
    return ","


def detect_text_encoding(path: Path) -> str:
    raw = path.read_bytes()[:262144]
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "utf-16-le", "utf-16-be", "cp1254", "cp1252", "latin-1"):
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def iter_text_lines(path: Path, encoding: str):
    with path.open("r", encoding=encoding, errors="replace") as handle:
        for line in handle:
            stripped = line.strip()
            if stripped and not stripped.startswith("#"):
                yield stripped


def parse_delimited_line(line: str, delimiter: str, headers: list[str]) -> dict:
    parts = split_line(line, delimiter)
    if headers:
        raw = {
            headers[index]: parts[index] if index < len(parts) else ""
            for index in range(len(headers))
        }
        return map_row(raw)
    return map_row_positional(parts)


def phone_index_fast(record: dict) -> str:
    phone = record.get("phone", "")
    if not phone:
        return ""
    keys = phone_keys(phone)
    return "|".join(sorted(keys, key=len, reverse=True)) if keys else ""


def tune_sqlite_for_bulk(conn: sqlite3.Connection) -> None:
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=OFF")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA cache_size=-500000")


def restore_sqlite_settings(conn: sqlite3.Connection) -> None:
    conn.execute("PRAGMA synchronous=NORMAL")


def read_delimited_rows(path: Path, suffix: str = ".csv") -> list[dict]:
    return list(iter_delimited_records(path, suffix))


def read_csv_rows(path: Path, suffix: str = ".csv") -> list[dict]:
    return read_delimited_rows(path, suffix)


def insert_record_batch(conn: sqlite3.Connection, batch: list[tuple]) -> None:
    if batch:
        conn.executemany(INSERT_SQL, batch)
        batch.clear()


def stream_records(conn: sqlite3.Connection, path: Path, records) -> int:
    batch: list[tuple] = []
    total = 0
    progress_every = PROGRESS_EVERY
    if path.stat().st_size >= LARGE_FILE_BYTES:
        progress_every = 100000

    for record in records:
        if not row_is_valid(record):
            continue
        batch.append(record_to_row(record))
        total += 1
        if len(batch) >= BATCH_SIZE:
            insert_record_batch(conn, batch)
        if total % progress_every == 0:
            print(f"  Indexed {total} rows from {path.name}...", flush=True)

    insert_record_batch(conn, batch)
    return total


def iter_xlsx_records(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    headers: list[str] = []
    for row in rows:
        headers = [clean_header(cell) for cell in row]
        if any(headers):
            break

    if not any(headers):
        workbook.close()
        return

    parsed = 0
    for row in rows:
        raw = {}
        for index, header in enumerate(headers):
            if not header or index >= len(row):
                continue
            raw[header] = row[index]
        parsed += 1
        if parsed % PROGRESS_EVERY == 0:
            print(f"  Indexed {parsed} rows from {path.name}...", flush=True)
        yield map_row(raw)

    workbook.close()


def iter_delimited_records(path: Path, suffix: str = ".csv"):
    encoding = detect_text_encoding(path)
    line_iter = iter_text_lines(path, encoding)
    sample = list(itertools.islice(line_iter, 100))

    if not sample:
        print(f"  {path.name} is empty", flush=True)
        return

    delimiter = choose_best_delimiter(sample, suffix)
    print(f"  Delimiter for {path.name}: {repr(delimiter)}", flush=True)

    first_parts = split_line(sample[0], delimiter)
    has_header = line_is_header(first_parts)
    headers = [clean_header(part) for part in first_parts] if has_header else []

    if has_header:
        print(f"  Headers for {path.name}: {headers[:12]}", flush=True)
    else:
        print(f"  No header row in {path.name}, using positional columns", flush=True)

    progress_every = PROGRESS_EVERY
    if path.stat().st_size >= LARGE_FILE_BYTES:
        progress_every = 100000
        print(f"  Large file detected ({path.stat().st_size // (1024 * 1024)} MB), streaming...", flush=True)

    valid = 0
    start = 1 if has_header else 0

    for line in itertools.chain(sample[start:], line_iter):
        record = parse_delimited_line(line, delimiter, headers)
        if not row_is_valid(record):
            continue

        valid += 1
        if valid <= 3:
            print(
                f"  Sample row {valid}: {record.get('first_name', '')} {record.get('last_name', '')} {record.get('phone', '') or record.get('identity_number', '')}",
                flush=True,
            )
        if valid % progress_every == 0:
            print(f"  Indexed {valid} rows from {path.name}...", flush=True)
        yield record

    if valid == 0:
        print(f"  No valid rows loaded from {path.name}. First line: {sample[0][:200]}", flush=True)


def read_xlsx_rows(path: Path) -> list[dict]:
    return list(iter_xlsx_records(path))


def iter_extracted_data_files(root: Path):
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("._"):
            continue
        if path.suffix.lower() in DATA_SUFFIXES:
            yield path


def load_records_from_extracted(root: Path, source_label: str) -> list[dict]:
    records: list[dict] = []
    found_any = False

    for file_path in iter_extracted_data_files(root):
        found_any = True
        try:
            file_records = load_file_records(file_path)
            records.extend(file_records)
            print(
                f"  Loaded {len(file_records)} records from {file_path.name}",
                flush=True,
            )
        except Exception as error:
            print(f"  Failed {file_path.name} in {source_label}: {error}", flush=True)

    if not found_any:
        names = [path.name for path in root.rglob("*") if path.is_file()]
        preview = ", ".join(names[:20]) or "no files"
        print(f"  No supported files in {source_label}. Found: {preview}", flush=True)

    return records


def read_7z_py7zr(path: Path, tmp_path: Path) -> None:
    import py7zr

    with py7zr.SevenZipFile(path, mode="r") as archive:
        names = [name.replace("\\", "/") for name in archive.getnames() if not name.endswith("/")]
        print(f"  7z contents: {', '.join(names[:20])}", flush=True)
        archive.extractall(path=tmp_path)


def read_7z_system(path: Path, tmp_path: Path) -> bool:
    command = shutil.which("7z") or shutil.which("7za")
    if not command:
        return False

    result = subprocess.run(
        [command, "x", "-y", f"-o{tmp_path}", str(path)],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"  7z command failed: {result.stderr.strip() or result.stdout.strip()}", flush=True)
        return False
    return True


def read_7z_records(path: Path) -> list[dict]:
    records: list[dict] = []
    py7zr_ok = False

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        py7zr_ok = False
        try:
            read_7z_py7zr(path, tmp_path)
            py7zr_ok = True
            records = load_records_from_extracted(tmp_path, path.name)
        except Exception as error:
            print(f"  py7zr failed on {path.name}: {error}", flush=True)

    if records:
        return records

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        if read_7z_system(path, tmp_path):
            records = load_records_from_extracted(tmp_path, path.name)
            if records:
                return records

    if not py7zr_ok:
        raise RuntimeError(
            f"Could not read {path.name}. Make sure p7zip-full is installed and the archive is not password protected."
        )

    return records


def free_disk_bytes(path: Path | None = None) -> int:
    target = path or BASE_DIR
    return shutil.disk_usage(target).free


def db_file_path(conn: sqlite3.Connection) -> Path | None:
    row = conn.execute("PRAGMA database_list").fetchone()
    if row and row[2]:
        return Path(row[2])
    return None


def count_text_lines(path: Path) -> int:
    wc = shutil.which("wc")
    if wc:
        result = subprocess.run(
            [wc, "-l", str(path)],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode == 0 and result.stdout.strip():
            return int(result.stdout.split()[0])

    total = 0
    with path.open("rb", buffering=16 * 1024 * 1024) as handle:
        while True:
            chunk = handle.read(16 * 1024 * 1024)
            if not chunk:
                break
            total += chunk.count(b"\n")
    return total


def count_xlsx_lines(path: Path) -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    total = 0
    for _ in sheet.iter_rows(values_only=True):
        total += 1
    workbook.close()
    return total


def count_db_lines(path: Path) -> int:
    if path.resolve() == INDEX_DB.resolve():
        return 0

    conn = sqlite3.connect(path)
    try:
        tables = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='people'"
        ).fetchall()
        if not tables:
            return 0
        return conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    finally:
        conn.close()


def file_has_header(path: Path) -> bool:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        return True
    if suffix != ".txt":
        return False

    with path.open("r", encoding="utf-8", errors="replace") as handle:
        first = handle.readline().strip()
    if not first:
        return False

    delimiter = "\t" if "\t" in first else "|" if "|" in first else ","
    parts = [part.strip() for part in first.split(delimiter)]
    return line_is_header(parts)


def count_file_lines(path: Path) -> int | None:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv", ".tsv"}:
        return count_text_lines(path)
    if suffix in {".xlsx", ".xlsm"}:
        return count_xlsx_lines(path)
    if suffix == ".db":
        return count_db_lines(path)
    return None


def file_stats(path: Path) -> dict:
    stat = path.stat()
    suffix = path.suffix.lower()
    lines = count_file_lines(path)
    data_lines = None

    if lines is not None:
        data_lines = max(lines - 1, 0) if file_has_header(path) and lines > 0 else lines

    return {
        "file": path.name,
        "type": suffix.lstrip(".") or "unknown",
        "size_bytes": stat.st_size,
        "size_mb": round(stat.st_size / (1024 * 1024), 2),
        "lines": lines,
        "data_lines": data_lines,
    }


def collect_stats(count_lines: bool = False) -> dict:
    files = source_files()
    indexed_meta: dict[str, dict] = {}
    if INDEX_DB.exists():
        try:
            conn = sqlite3.connect(INDEX_DB)
            try:
                indexed_meta = load_indexed_sources(conn)
            finally:
                conn.close()
        except sqlite3.Error:
            indexed_meta = {}

    file_rows = []
    for path in files:
        stat = path.stat()
        suffix = path.suffix.lower()
        meta = indexed_meta.get(path.name, {})
        row = {
            "file": path.name,
            "type": suffix.lstrip(".") or "unknown",
            "size_bytes": stat.st_size,
            "size_mb": round(stat.st_size / (1024 * 1024), 2),
            "lines": None,
            "data_lines": None,
            "indexed_records": meta.get("records"),
            "needs_import": file_needs_import(path, indexed_meta) if indexed_meta else True,
        }
        if count_lines:
            lines = count_file_lines(path)
            row["lines"] = lines
            if lines is not None:
                row["data_lines"] = (
                    max(lines - 1, 0) if file_has_header(path) and lines > 0 else lines
                )
        file_rows.append(row)

    line_counts = [row["lines"] for row in file_rows if isinstance(row["lines"], int)]

    stats = {
        "files": len(file_rows),
        "total_lines": sum(line_counts) if count_lines else None,
        "total_size_bytes": sum(row["size_bytes"] for row in file_rows),
        "total_size_mb": round(sum(row["size_bytes"] for row in file_rows) / (1024 * 1024), 2),
        "free_disk_mb": free_disk_bytes() // (1024 * 1024),
        "index_ready": INDEX_READY.is_set(),
        "index_building": INDEX_BUILDING.is_set(),
        "auto_rebuild": AUTO_REBUILD,
        "indexed_records": (
            count_records()
            if INDEX_READY.is_set() and INDEX_DB.exists() and not INDEX_BUILDING.is_set()
            else 0
        ),
        "sources": file_rows,
    }

    if INDEX_ERROR:
        stats["index_error"] = INDEX_ERROR

    return stats


def source_files() -> list[Path]:
    files: list[Path] = []
    if not DATABASE_DIR.exists():
        return files
    for path in sorted(DATABASE_DIR.iterdir()):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix in DATA_SUFFIXES or suffix in ARCHIVE_SUFFIXES:
            files.append(path)
    return files


def connect_index(timeout: float = 60.0) -> sqlite3.Connection:
    conn = sqlite3.connect(INDEX_DB, check_same_thread=False, timeout=timeout)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=60000")
    return conn


def ensure_index_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL DEFAULT '',
            last_name TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            email TEXT NOT NULL DEFAULT '',
            identity_number TEXT NOT NULL DEFAULT '',
            city TEXT NOT NULL DEFAULT '',
            country TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            extra_json TEXT NOT NULL DEFAULT '{}',
            first_name_n TEXT NOT NULL DEFAULT '',
            last_name_n TEXT NOT NULL DEFAULT '',
            phone_n TEXT NOT NULL DEFAULT '',
            email_n TEXT NOT NULL DEFAULT '',
            identity_number_n TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta (key, value) VALUES ('schema_version', ?)",
        (str(SCHEMA_VERSION),),
    )


def ensure_index_schema_if_needed(conn: sqlite3.Connection) -> None:
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='people'"
    ).fetchone()
    if not tables:
        ensure_index_schema(conn)
        return
    stored = conn.execute(
        "SELECT value FROM meta WHERE key='schema_version'"
    ).fetchone()
    if not stored or int(stored[0]) != SCHEMA_VERSION:
        ensure_index_schema(conn)


def record_to_row(record: dict) -> tuple:
    extra = record.get("extra") or {}
    return (
        record["first_name"],
        record["last_name"],
        record["phone"],
        record["email"],
        record["identity_number"],
        record["city"],
        record["country"],
        record["notes"],
        "{}" if not extra else json.dumps(extra, ensure_ascii=False),
        norm_text(record["first_name"]),
        norm_text(record["last_name"]),
        phone_index_fast(record) if not extra else collect_phone_keys(record),
        norm_email(record["email"]),
        norm_text(record["identity_number"]),
    )


def insert_records(conn: sqlite3.Connection, records: list[dict]) -> None:
    batch: list[tuple] = []
    for record in records:
        batch.append(record_to_row(record))
        if len(batch) >= BATCH_SIZE:
            conn.executemany(INSERT_SQL, batch)
            batch.clear()
    if batch:
        conn.executemany(INSERT_SQL, batch)


def sql_quote(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def map_table_columns(column_names: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for column in column_names:
        field = COLUMN_MAP.get(clean_header(column))
        if field in IMPORT_FIELDS and field not in mapping:
            mapping[field] = column
    return mapping


def table_mapping_score(mapping: dict[str, str]) -> int:
    score = len(mapping)
    if "phone" in mapping or "identity_number" in mapping:
        score += 3
    if "email" in mapping:
        score += 2
    if "first_name" in mapping and "last_name" in mapping:
        score += 2
    return score


def list_data_tables(conn: sqlite3.Connection) -> list[str]:
    tables = [
        row[0]
        for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
    ]

    def sort_key(name: str) -> tuple[int, int, str]:
        clean = clean_header(name)
        if clean == "people":
            return (0, 0, name)
        if clean in TABLE_NAME_HINTS:
            return (1, 0, name)
        return (2, 0, name)

    return sorted(tables, key=sort_key)


def load_indexed_sources(conn: sqlite3.Connection) -> dict[str, dict]:
    row = conn.execute(
        "SELECT value FROM meta WHERE key = ?", (SOURCE_META_KEY,)
    ).fetchone()
    if not row:
        return {}
    try:
        raw = json.loads(row[0] or "{}")
        normalized: dict[str, dict] = {}
        for key, value in raw.items():
            if isinstance(value, dict):
                normalized[str(key)] = {
                    "mtime": float(value.get("mtime", 0)),
                    "records": int(value.get("records", 0)),
                }
            else:
                normalized[str(key)] = {"mtime": float(value), "records": 0}
        return normalized
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}


def save_indexed_sources(conn: sqlite3.Connection, sources: dict[str, dict]) -> None:
    conn.execute(
        "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
        (SOURCE_META_KEY, json.dumps(sources, ensure_ascii=False)),
    )


def source_is_importable(path: Path) -> bool:
    return path.suffix.lower() in {".sql", ".db"}


def file_needs_import(path: Path, indexed: dict[str, dict]) -> bool:
    mtime = path.stat().st_mtime
    entry = indexed.get(path.name)

    if entry is None:
        return True

    if source_is_importable(path):
        if entry.get("records", 0) <= 0:
            return True
        return entry.get("mtime") != mtime

    if entry.get("records", 0) == -1 and entry.get("mtime") == mtime:
        return False
    if entry.get("mtime") != mtime:
        return True
    return entry.get("records", 0) <= 0


def import_database_path(conn: sqlite3.Connection, path: Path, label: str) -> int:
    if path.resolve() == INDEX_DB.resolve():
        return 0

    before = conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    source = sqlite3.connect(path)
    source.row_factory = sqlite3.Row
    try:
        imported_any = False
        for table in list_data_tables(source):
            columns = [
                row[1] for row in source.execute(f"PRAGMA table_info({sql_quote(table)})")
            ]
            mapping = map_table_columns(columns)
            if table_mapping_score(mapping) < 2:
                continue

            select_parts = []
            for field in IMPORT_FIELDS:
                column = mapping.get(field)
                if column:
                    select_parts.append(f"{sql_quote(column)} AS {field}")
                else:
                    select_parts.append(f"'' AS {field}")

            batch: list[tuple] = []
            table_count = 0
            query = f"SELECT {', '.join(select_parts)} FROM {sql_quote(table)}"
            for row in source.execute(query):
                record = {field: str(row[field] or "").strip() for field in IMPORT_FIELDS}
                record["extra"] = {}
                if not row_is_valid(record):
                    continue
                batch.append(record_to_row(record))
                table_count += 1
                if len(batch) >= BATCH_SIZE:
                    conn.executemany(INSERT_SQL, batch)
                    batch.clear()

            if batch:
                conn.executemany(INSERT_SQL, batch)

            if table_count:
                imported_any = True
                print(f"  Imported {table_count} rows from {label} table {table}", flush=True)

        if not imported_any:
            tables = list_data_tables(source)
            raise RuntimeError(
                f"No searchable columns found in {label}. "
                f"Tables: {', '.join(tables[:10]) or 'none'}"
            )
    finally:
        source.close()

    after = conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    return after - before


def is_postgresql_dump(path: Path) -> bool:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        sample = handle.read(16384)
    lowered = sample.lower()
    return (
        "postgresql database dump" in lowered
        or ("copy " in lowered and "from stdin" in lowered)
        or "pg_catalog" in lowered
    )


def pg_clean_value(value: str) -> str:
    text = str(value or "").strip()
    if text in PG_NULLS:
        return ""
    return text


def pg_copy_record(columns: list[str], parts: list[str]) -> dict | None:
    raw: dict[str, str] = {}
    for index, column in enumerate(columns):
        raw[column] = pg_clean_value(parts[index] if index < len(parts) else "")

    record = {field: "" for field in IMPORT_FIELDS}
    record["extra"] = {}

    for column, value in raw.items():
        if not value:
            continue
        header = clean_header(column)
        field = COLUMN_MAP.get(header)
        if field in IMPORT_FIELDS and not record[field]:
            record[field] = value
        else:
            record["extra"][header] = value

    if not record["city"]:
        for key in ("address_city", "birth_city", "id_registration_city"):
            city = record["extra"].get(key, "")
            if city:
                record["city"] = city
                break

    note_bits = []
    for key in (
        "address_district", "address_neighborhood", "street_address",
        "door_or_entrance_number", "mother_first", "father_first",
        "date_of_birth", "gender", "misc",
    ):
        value = record["extra"].get(key, "")
        if value:
            note_bits.append(value)
    if note_bits and not record["notes"]:
        record["notes"] = " | ".join(note_bits)

    if not row_is_valid(record):
        return None
    return record


def parse_pg_copy_header(line: str) -> tuple[str, list[str]] | None:
    stripped = line.strip()
    lowered = stripped.lower()
    if not lowered.startswith("copy ") or "from stdin" not in lowered:
        return None

    match = PG_COPY_RE.match(stripped)
    if match:
        table = match.group(1) or match.group(2) or "unknown"
        columns = [part.strip().strip('"') for part in match.group(3).split(",")]
        return table, columns

    col_start = stripped.find("(")
    col_end = stripped.rfind(")")
    if col_start == -1 or col_end <= col_start:
        return None

    columns = [part.strip().strip('"') for part in stripped[col_start + 1 : col_end].split(",")]
    head = stripped[:col_start].strip()
    table_token = head[4:].strip().split()[-1]
    table = table_token.split(".")[-1].strip('"')
    return table, columns


def iter_postgresql_copy_rows(path: Path):
    columns: list[str] = []
    table = ""
    in_copy = False

    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            stripped = line.strip()
            if not in_copy:
                parsed = parse_pg_copy_header(line)
                if parsed:
                    table, columns = parsed
                    in_copy = True
                    print(f"  PostgreSQL COPY {table} ({len(columns)} columns)", flush=True)
                continue

            if stripped == "\\." or stripped.startswith("\\."):
                in_copy = False
                columns = []
                table = ""
                continue

            if not columns:
                continue

            parts = line.rstrip("\n\r").split("\t")
            record = pg_copy_record(columns, parts)
            if record is not None:
                yield table, record


def import_postgresql_dump(conn: sqlite3.Connection, path: Path) -> int:
    print(f"  PostgreSQL dump detected in {path.name}", flush=True)
    before = conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    batch: list[tuple] = []
    total = 0
    started = time.perf_counter()

    for _table, record in iter_postgresql_copy_rows(path):
        batch.append(record_to_row(record))
        total += 1
        if len(batch) >= BATCH_SIZE:
            conn.executemany(INSERT_SQL, batch)
            batch.clear()
        if total % PG_PROGRESS_EVERY == 0:
            elapsed = max(time.perf_counter() - started, 0.001)
            print(f"  Imported {total} PostgreSQL rows ({int(total / elapsed)}/sec)...", flush=True)

    if batch:
        conn.executemany(INSERT_SQL, batch)

    if total == 0:
        raise RuntimeError(f"No COPY data rows found in PostgreSQL dump {path.name}")

    after = conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
    print(f"  PostgreSQL import done: {after - before} records from {path.name}", flush=True)
    return after - before


def import_sql_file(conn: sqlite3.Connection, path: Path) -> int:
    if is_postgresql_dump(path):
        return import_postgresql_dump(conn, path)

    db_path = db_file_path(conn)
    if not db_path:
        raise RuntimeError("sql import needs a file-backed database")

    print(f"  SQLite SQL import for {path.name}...", flush=True)

    with tempfile.TemporaryDirectory() as tmp_dir:
        import_db = Path(tmp_dir) / "import.db"
        script = f".read {path.resolve()}"
        result = subprocess.run(
            ["sqlite3", str(import_db)],
            input=script,
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "sql import failed")
        return import_database_path(conn, import_db, path.name)


def index_db_file(conn: sqlite3.Connection, path: Path) -> int:
    print(f"  Importing database {path.name}...", flush=True)
    return import_database_path(conn, path, path.name)


def read_db_rows(path: Path) -> list[dict]:
    if path.resolve() == INDEX_DB.resolve():
        return []

    rows: list[dict] = []
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        for table in list_data_tables(conn):
            columns = [row[1] for row in conn.execute(f"PRAGMA table_info({sql_quote(table)})")]
            mapping = map_table_columns(columns)
            if table_mapping_score(mapping) < 2:
                continue
            select_parts = []
            for field in IMPORT_FIELDS:
                column = mapping.get(field)
                if column:
                    select_parts.append(f"{sql_quote(column)} AS {field}")
                else:
                    select_parts.append(f"'' AS {field}")
            query = f"SELECT {', '.join(select_parts)} FROM {sql_quote(table)}"
            for row in conn.execute(query):
                record = {field: str(row[field] or "").strip() for field in IMPORT_FIELDS}
                record["extra"] = {}
                if row_is_valid(record):
                    rows.append(record)
    finally:
        conn.close()
    return rows


def load_file_records(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        records = list(iter_xlsx_records(path))
    elif suffix in {".csv", ".tsv", ".txt"}:
        records = list(iter_delimited_records(path, suffix))
    elif suffix == ".db":
        records = read_db_rows(path)
    elif suffix == ".sql":
        return []
    elif suffix == ".7z":
        records = read_7z_records(path)
    else:
        records = []
    return [record for record in records if row_is_valid(record)]


def index_file(conn: sqlite3.Connection, path: Path) -> int:
    suffix = path.suffix.lower()
    print(f"Loading {path.name}...", flush=True)

    if suffix == ".sql":
        return import_sql_file(conn, path)
    if suffix == ".db":
        return index_db_file(conn, path)
    if suffix in {".xlsx", ".xlsm"}:
        return stream_records(conn, path, iter_xlsx_records(path))
    if suffix in {".csv", ".tsv", ".txt"}:
        return stream_records(conn, path, iter_delimited_records(path, suffix))

    records = load_file_records(path)
    insert_records(conn, records)
    return len(records)


def index_usable() -> bool:
    if not INDEX_DB.exists():
        return False

    conn = sqlite3.connect(INDEX_DB)
    try:
        tables = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='people'"
        ).fetchone()
        if not tables:
            return False
        columns = {row[1] for row in conn.execute("PRAGMA table_info(people)")}
        required = {
            "first_name", "last_name", "phone", "email", "identity_number",
            "city", "country", "notes", "extra_json", "phone_n",
        }
        return required.issubset(columns)
    except sqlite3.Error:
        return False
    finally:
        conn.close()


def index_schema_ok() -> bool:
    if not index_usable():
        return False

    conn = sqlite3.connect(INDEX_DB)
    try:
        version = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='meta'"
        ).fetchone()
        if not version:
            return False
        stored = conn.execute("SELECT value FROM meta WHERE key='schema_version'").fetchone()
        if not stored or int(stored[0]) != SCHEMA_VERSION:
            return False

        return True
    except (sqlite3.Error, ValueError, TypeError):
        return False
    finally:
        conn.close()


def index_is_stale() -> bool:
    if not INDEX_DB.exists():
        return True
    if not index_schema_ok():
        return True
    index_mtime = INDEX_DB.stat().st_mtime
    for path in source_files():
        if path.stat().st_mtime > index_mtime:
            return True
    return False


def bootstrap_source_meta(conn: sqlite3.Connection) -> dict[str, dict]:
    indexed = load_indexed_sources(conn)
    if indexed:
        return indexed

    meta: dict[str, dict] = {}
    for path in source_files():
        if source_is_importable(path):
            continue
        meta[path.name] = {"mtime": path.stat().st_mtime, "records": -1}

    if meta:
        save_indexed_sources(conn, meta)
        print(
            f"Marked {len(meta)} existing data file(s). SQL/DB files will import on startup.",
            flush=True,
        )
    return load_indexed_sources(conn)


def index_pending_sources(conn: sqlite3.Connection, sql_db_only: bool = True) -> int:
    indexed = bootstrap_source_meta(conn)
    total_added = 0
    updated = dict(indexed)

    for path in source_files():
        if sql_db_only and not source_is_importable(path):
            continue
        if not file_needs_import(path, indexed):
            continue
        try:
            added = index_file(conn, path)
            updated[path.name] = {"mtime": path.stat().st_mtime, "records": added}
            total_added += added
            print(f"Indexed {added} records from {path.name}", flush=True)
        except Exception as error:
            print(f"Failed to index {path.name}: {error}", flush=True)
            updated[path.name] = {"mtime": path.stat().st_mtime, "records": 0}

    save_indexed_sources(conn, updated)
    return total_added


def index_sql_db_sources(conn: sqlite3.Connection) -> int:
    indexed = load_indexed_sources(conn)
    for path in source_files():
        if source_is_importable(path):
            indexed.pop(path.name, None)
    save_indexed_sources(conn, indexed)
    conn.commit()
    return index_pending_sources(conn, sql_db_only=True)


def ensure_index(force: bool = False) -> None:
    with INDEX_LOCK:
        if not force and index_usable():
            if AUTO_REBUILD and index_is_stale():
                print("AUTO_REBUILD=1: rebuilding stale index...", flush=True)
            else:
                conn = sqlite3.connect(INDEX_DB)
                try:
                    tune_sqlite_for_bulk(conn)
                    ensure_index_schema(conn)
                    added = index_pending_sources(conn)
                    conn.commit()
                    restore_sqlite_settings(conn)
                    records = count_records()
                    if added:
                        print(f"Added {added} records from new SQL/DB files", flush=True)
                    print(
                        f"Search index ready ({records} records). "
                        "Full rebuild: GET /api/rebuild",
                        flush=True,
                    )
                finally:
                    conn.close()
                return

        info = rebuild_index()
        print(
            f"Loaded {info['records']} records from {len(info['sources'])} file(s)",
            flush=True,
        )


def rebuild_index_async(force: bool = True) -> bool:
    if INDEX_BUILDING.is_set():
        return False

    def worker() -> None:
        global INDEX_ERROR
        INDEX_BUILDING.set()
        INDEX_READY.clear()
        INDEX_ERROR = None
        try:
            ensure_index(force=force)
        except Exception as error:
            INDEX_ERROR = str(error)
            print(f"Index build failed: {error}", flush=True)
        finally:
            INDEX_BUILDING.clear()
            INDEX_READY.set()

    threading.Thread(target=worker, daemon=True).start()
    return True


def build_index_background() -> None:
    global INDEX_ERROR
    INDEX_BUILDING.set()
    try:
        with INDEX_LOCK:
            if INDEX_DB.exists() and index_usable():
                ensure_index(force=False)
            else:
                print("No index found, building once in background...", flush=True)
                ensure_index(force=True)
    except Exception as error:
        INDEX_ERROR = str(error)
        print(f"Index build failed: {error}", flush=True)
    finally:
        INDEX_BUILDING.clear()
        INDEX_READY.set()


def wait_for_index() -> None:
    if not INDEX_READY.wait(timeout=None):
        raise HTTPException(status_code=503, detail="Index is still building, try again in a few minutes")
    if INDEX_BUILDING.is_set():
        raise HTTPException(status_code=503, detail="Index rebuild in progress, try again in a few minutes")
    if INDEX_ERROR and not (index_usable() and count_records() > 0):
        raise HTTPException(status_code=500, detail=INDEX_ERROR)


def rebuild_index() -> dict:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    temp_db = BASE_DIR / ".search_index.building.db"
    if temp_db.exists():
        temp_db.unlink()

    sources = sorted(source_files(), key=lambda item: (0 if item.suffix.lower() == ".sql" else 1, item.stat().st_size))
    loaded_files = []
    file_counts: dict[str, int] = {}
    total_records = 0

    conn = sqlite3.connect(temp_db)
    try:
        tune_sqlite_for_bulk(conn)
        ensure_index_schema(conn)
        for path in sources:
            try:
                conn.execute("BEGIN IMMEDIATE")
                count = index_file(conn, path)
                conn.commit()
            except Exception as error:
                conn.execute("ROLLBACK")
                print(f"Failed to load {path.name}: {error}", flush=True)
                file_counts[path.name] = 0
                continue
            file_counts[path.name] = count
            if count:
                loaded_files.append(path.name)
                total_records += count
                print(f"Loaded {count} records from {path.name}", flush=True)
        restore_sqlite_settings(conn)
        save_indexed_sources(
            conn,
            {
                path.name: {"mtime": path.stat().st_mtime, "records": file_counts.get(path.name, 0)}
                for path in sources
            },
        )
        conn.commit()
    finally:
        conn.close()

    if INDEX_DB.exists():
        INDEX_DB.unlink()
    temp_db.replace(INDEX_DB)

    final_conn = sqlite3.connect(INDEX_DB)
    try:
        save_indexed_sources(
            final_conn,
            {
                path.name: {"mtime": path.stat().st_mtime, "records": file_counts.get(path.name, 0)}
                for path in sources
            },
        )
        final_conn.commit()
    finally:
        final_conn.close()

    return {
        "sources": loaded_files,
        "records": total_records,
    }


def count_records() -> int:
    if not INDEX_DB.exists():
        return 0
    try:
        conn = connect_index()
        try:
            return conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
        finally:
            conn.close()
    except sqlite3.Error:
        return 0


def format_result(row: sqlite3.Row) -> dict:
    extra = {}
    try:
        extra = json.loads(row["extra_json"] or "{}")
    except json.JSONDecodeError:
        extra = {}

    result = {
        "first_name": row["first_name"],
        "last_name": row["last_name"],
        "full_name": f"{row['first_name']} {row['last_name']}".strip(),
        "phone": row["phone"],
        "email": row["email"],
        "identity_number": row["identity_number"],
        "city": row["city"],
        "country": row["country"],
        "notes": row["notes"],
    }
    if extra:
        result["extra"] = extra
    return result


def parse_search_query(q: str | None) -> dict:
    text = re.sub(r"\s+", " ", (q or "").strip())
    if not text:
        return {}

    if "@" in text and "." in text.split("@", 1)[-1]:
        return {"type": "email", "email": text, "q": text}

    digits = phone_digits(text)
    compact = re.sub(r"\s+", "", text)

    if digits:
        if looks_like_tc(digits):
            return {"type": "identity_number", "identity_number": digits, "q": text}

        if looks_like_phone_digits(digits):
            return {"type": "phone", "phone": text, "q": text}

        digit_ratio = len(digits) / max(len(compact), 1)
        if len(digits) >= 7 and digit_ratio >= 0.7:
            return {"type": "phone", "phone": text, "q": text}

    parts = text.split(" ")
    if len(parts) == 1:
        return {"type": "name", "name": text, "q": text}

    return {
        "type": "name",
        "q": text,
        "first_name": parts[0],
        "last_name": " ".join(parts[1:]),
        "full_name": text,
    }


def search(q: str | None = None) -> tuple[list[dict], dict]:
    wait_for_index()

    parsed = parse_search_query(q)
    if not parsed:
        raise ValueError("Send q with a name, phone, or email")

    phone = parsed.get("phone")
    email = parsed.get("email")
    first_name = parsed.get("first_name")
    last_name = parsed.get("last_name")
    name = parsed.get("name")
    identity_number = parsed.get("identity_number")

    clauses = []
    params: list = []

    if phone:
        variants = phone_search_variants(phone)
        if not variants:
            raise ValueError("Invalid phone number")

        phone_parts = []
        for variant in sorted(variants, key=len, reverse=True):
            phone_parts.append("phone_n LIKE ?")
            params.append(f"%{variant}%")
            phone_parts.append(f"{PHONE_DIGITS_SQL} LIKE ?")
            params.append(f"%{variant}%")
            phone_parts.append(f"{IDENTITY_DIGITS_SQL} LIKE ?")
            params.append(f"%{variant}%")
            phone_parts.append("notes LIKE ?")
            params.append(f"%{variant}%")
            phone_parts.append("extra_json LIKE ?")
            params.append(f"%{variant}%")

        clauses.append(f"({' OR '.join(phone_parts)})")

    if email:
        clauses.append("email_n LIKE ?")
        params.append(f"%{norm_email(email)}%")

    if first_name and last_name:
        clauses.append("first_name_n LIKE ?")
        params.append(f"%{norm_text(first_name)}%")
        clauses.append("last_name_n LIKE ?")
        params.append(f"%{norm_text(last_name)}%")
    elif name:
        needle = norm_text(name)
        clauses.append(
            "(first_name_n LIKE ? OR last_name_n LIKE ? OR (first_name_n || ' ' || last_name_n) LIKE ?)"
        )
        params.extend([f"%{needle}%", f"%{needle}%", f"%{needle}%"])

    if identity_number:
        id_digits = phone_digits(identity_number)
        id_parts = ["identity_number_n LIKE ?"]
        id_params = [f"%{norm_text(identity_number)}%"]
        if id_digits:
            id_parts.append(f"{IDENTITY_DIGITS_SQL} LIKE ?")
            id_params.append(f"%{id_digits}%")
        clauses.append(f"({' OR '.join(id_parts)})")
        params.extend(id_params)

    query_sql = f"""
        SELECT
            first_name, last_name, phone, email, identity_number,
            city, country, notes, extra_json
        FROM people
        WHERE {' AND '.join(clauses)}
    """

    with connect_index() as conn:
        ensure_index_schema_if_needed(conn)
        rows = conn.execute(query_sql, params).fetchall()

    results = [format_result(row) for row in rows]
    return results, parsed


def raw_json(**data) -> Response:
    payload = {"credit": CREDIT, "version": API_VERSION, **data}
    return Response(
        content=json.dumps(payload, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
    )


@app.get("/api/stats")
def api_stats(
    count_lines: bool = Query(default=False, description="Scan files to count lines (slow on big files)"),
) -> Response:
    started = time.perf_counter()
    return raw_json(
        ok=True,
        stats=collect_stats(count_lines=count_lines),
        ms=round((time.perf_counter() - started) * 1000, 2),
    )


@app.get("/api/reimport")
def api_reimport() -> Response:
    if not INDEX_DB.exists() or not index_usable():
        rebuild_index_async(force=True)
        return raw_json(
            ok=True,
            rebuilding=True,
            message="No index yet. Full rebuild started.",
        )

    if INDEX_BUILDING.is_set():
        return raw_json(ok=True, rebuilding=True, message="Import already running. Wait for it to finish.")

    def worker() -> None:
        global INDEX_ERROR
        INDEX_BUILDING.set()
        INDEX_ERROR = None
        try:
            with INDEX_LOCK:
                conn = connect_index()
                try:
                    tune_sqlite_for_bulk(conn)
                    ensure_index_schema_if_needed(conn)
                    added = index_sql_db_sources(conn)
                    conn.commit()
                    restore_sqlite_settings(conn)
                    print(f"Reimport added {added} records from SQL/DB files", flush=True)
                finally:
                    conn.close()
        except Exception as error:
            INDEX_ERROR = str(error)
            print(f"Reimport failed: {error}", flush=True)
        finally:
            INDEX_BUILDING.clear()
            INDEX_READY.set()

    threading.Thread(target=worker, daemon=True).start()
    return raw_json(
        ok=True,
        rebuilding=True,
        message="Reimporting SQL/DB files in background. Check /api/stats for indexed_records.",
    )


@app.get("/api/rebuild")
def api_rebuild() -> Response:
    if INDEX_BUILDING.is_set():
        return raw_json(ok=True, rebuilding=True, message="Rebuild already running")
    rebuild_index_async(force=True)
    return raw_json(
        ok=True,
        rebuilding=True,
        message="Rebuild started in background. Check /api for ready status.",
    )


@app.get("/api")
def api(
    q: str | None = Query(default=None, description="Name, phone, email, or ID number"),
) -> Response:
    started = time.perf_counter()

    if not q or not q.strip():
        if not INDEX_READY.is_set() or INDEX_BUILDING.is_set():
            return raw_json(
                ok=True,
                ready=False,
                status="indexing" if INDEX_BUILDING.is_set() else "starting",
                records=count_records() if INDEX_DB.exists() and not INDEX_BUILDING.is_set() else 0,
                free_disk_mb=free_disk_bytes() // (1024 * 1024),
                usage=API_USAGE,
            )
        return raw_json(
            ok=True,
            ready=True,
            status="ready",
            records=count_records(),
            usage=API_USAGE,
        )

    try:
        results, query = search(q)
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    except HTTPException:
        raise
    except sqlite3.Error as error:
        raise HTTPException(status_code=500, detail=str(error)) from error
    except Exception as error:
        raise HTTPException(status_code=500, detail=str(error)) from error

    return raw_json(
        ok=True,
        ready=INDEX_READY.is_set(),
        success=True,
        query=query,
        found=len(results),
        returned=len(results),
        results=results,
        ms=round((time.perf_counter() - started) * 1000, 2),
    )


@app.get("/api/search")
def api_search_legacy(
    q: str | None = Query(default=None),
    phone: str | None = Query(default=None),
    email: str | None = Query(default=None),
    first_name: str | None = Query(default=None),
    last_name: str | None = Query(default=None),
    identity_number: str | None = Query(default=None),
) -> Response:
    if q:
        return api(q=q)
    if phone:
        return api(q=phone)
    if email:
        return api(q=email)
    if first_name and last_name:
        return api(q=f"{first_name} {last_name}")
    if first_name:
        return api(q=first_name)
    if last_name:
        return api(q=last_name)
    if identity_number:
        return api(q=identity_number)
    return api()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT)
